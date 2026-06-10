import os
import sys
import json
import logging
import threading
import datetime
import traceback
import re
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import openpyxl
from openpyxl.styles import Font, Alignment
from striprtf.striprtf import rtf_to_text
import google.generativeai as genai
import typing_extensions as typing

def apply_tnr_font(ws):
    tnr_font = Font(name='Times New Roman', size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = tnr_font


# Cấu hình log
LOG_FILE = "app_bien_muc_tg24h.log"
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8"
)

CONFIG_FILE = "config.json"

# --- Định nghĩa Schema cho Gemini API ---
class NewsItem(typing.TypedDict):
    id: str
    ten_file: str
    thoi_luong: str

class ListParseResult(typing.TypedDict):
    ngay_phat_song: str
    danh_sach_tin: list[NewsItem]

class CrewList(typing.TypedDict):
    chiu_trach_nhiem: str
    bien_tap: str
    bien_dich: str
    hien_dan: str
    dao_dien: str
    ky_thuat: str
    trang_diem: str

class RtfParseResult(typing.TypedDict):
    tieu_de: str
    nguoi_bien_dich: str
    noi_dung: list[str]

# --- App Chính ---
class BienMucApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tự Động Biên Mục Thế Giới 24H (TG24H)")
        self.root.geometry("800x600")
        self.root.resizable(True, True)

        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.api_key = tk.StringVar()
        self.model_name = tk.StringVar(value="gemini-1.5-flash") # Mặc định
        
        self.load_config()
        self.build_ui()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    self.api_key.set(cfg.get("api_key", ""))
                    self.model_name.set(cfg.get("model_name", "gemini-1.5-flash"))
            except Exception as e:
                logging.error(f"Lỗi đọc config: {e}")
        
        if not self.api_key.get() and "GEMINI_API_KEY" in os.environ:
            self.api_key.set(os.environ["GEMINI_API_KEY"])

    def save_config(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({
                    "api_key": self.api_key.get(),
                    "model_name": self.model_name.get()
                }, f, indent=4)
        except Exception as e:
            logging.error(f"Lỗi lưu config: {e}")

    def build_ui(self):
        # Frame trên: Chọn folder
        frame_top = ttk.LabelFrame(self.root, text="Cấu hình Đường dẫn", padding=(10, 10))
        frame_top.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_top, text="Thư mục Input:").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Entry(frame_top, textvariable=self.input_dir, width=60).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(frame_top, text="Chọn...", command=self.browse_input).grid(row=0, column=2, pady=5)

        ttk.Label(frame_top, text="Thư mục Output:").grid(row=1, column=0, sticky="w", pady=5)
        ttk.Entry(frame_top, textvariable=self.output_dir, width=60).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(frame_top, text="Chọn...", command=self.browse_output).grid(row=1, column=2, pady=5)
        ttk.Label(frame_top, text="(Mặc định tạo folder 'output' trong input nếu để trống)", font=("Arial", 8, "italic")).grid(row=2, column=1, sticky="w")

        # Frame giữa: Controls
        frame_mid = ttk.Frame(self.root, padding=(10, 5))
        frame_mid.pack(fill="x", padx=10)

        ttk.Button(frame_mid, text="⚙ Cài đặt API", command=self.open_settings).pack(side="left")
        
        self.btn_start = ttk.Button(frame_mid, text="▶ BẮT ĐẦU BIÊN MỤC", command=self.start_process, style="Accent.TButton")
        self.btn_start.pack(side="right")

        # Progress
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.root, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", padx=10, pady=10)

        # Log Text
        self.txt_log = scrolledtext.ScrolledText(self.root, state="disabled", wrap="word", font=("Consolas", 10))
        self.txt_log.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def browse_input(self):
        d = filedialog.askdirectory(title="Chọn thư mục Input chứa file XLSX và RTF")
        if d: self.input_dir.set(d)

    def browse_output(self):
        d = filedialog.askdirectory(title="Chọn thư mục Output")
        if d: self.output_dir.set(d)

    def log(self, message):
        logging.info(message)
        self.root.after(0, self._append_log, message)

    def _append_log(self, message):
        self.txt_log.config(state="normal")
        self.txt_log.insert(tk.END, f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state="disabled")

    def open_settings(self):
        top = tk.Toplevel(self.root)
        top.title("Cài đặt API")
        top.geometry("500x200")
        top.transient(self.root)
        top.grab_set()

        ttk.Label(top, text="Gemini API Key:").pack(anchor="w", padx=10, pady=(10, 0))
        ent_key = ttk.Entry(top, textvariable=self.api_key, show="*")
        ent_key.pack(fill="x", padx=10, pady=5)

        ttk.Label(top, text="Model Name (VD: gemini-1.5-flash, gemini-1.5-pro, gemma-4-26b-a4b-it):").pack(anchor="w", padx=10, pady=(10, 0))
        ent_model = ttk.Entry(top, textvariable=self.model_name)
        ent_model.pack(fill="x", padx=10, pady=5)

        def save():
            self.save_config()
            self.log("Đã lưu cấu hình API.")
            top.destroy()

        ttk.Button(top, text="Lưu & Đóng", command=save).pack(pady=15)

    def start_process(self):
        input_d = self.input_dir.get().strip()
        if not input_d or not os.path.isdir(input_d):
            messagebox.showerror("Lỗi", "Vui lòng chọn thư mục Input hợp lệ!")
            return
        
        if not self.api_key.get().strip():
            messagebox.showerror("Lỗi", "Vui lòng nhập Gemini API Key trong phần Cài đặt!")
            return

        self.btn_start.config(state="disabled")
        self.progress_var.set(0)
        self.txt_log.config(state="normal")
        self.txt_log.delete(1.0, tk.END)
        self.txt_log.config(state="disabled")
        
        thread = threading.Thread(target=self.process_thread)
        thread.daemon = True
        thread.start()

    def process_thread(self):
        try:
            self.log("=== BẮT ĐẦU TIẾN TRÌNH ===")
            genai.configure(api_key=self.api_key.get().strip())
            model_name = self.model_name.get().strip()
            
            self.log(f"Đang sử dụng model: {model_name}")
            
            input_d = self.input_dir.get().strip()
            output_d = self.output_dir.get().strip()
            if not output_d:
                output_d = os.path.join(input_d, "output")
            os.makedirs(output_d, exist_ok=True)
            self.log(f"Thư mục Output: {output_d}")

            # 1. Tìm file LIST
            list_files = [f for f in os.listdir(input_d) if f.startswith("BTTG24H_") and f.endswith(".xlsx")]
            if not list_files:
                raise Exception("Không tìm thấy file danh sách bắt đầu bằng 'BTTG24H_' và đuôi '.xlsx' trong thư mục Input.")
            list_file_path = os.path.join(input_d, list_files[0])
            self.log(f"Đã tìm thấy file danh sách: {list_files[0]}")

            # Đọc file LIST ra dạng text thô để gửi cho Gemini
            self.progress_var.set(10)
            self.log("Đọc dữ liệu file LIST...")
            wb_list = openpyxl.load_workbook(list_file_path, data_only=True)
            ws_list = wb_list.active
            
            list_text_data = []
            id_to_time = {}
            for row in ws_list.iter_rows(values_only=False):
                # Format time correctly from column F
                val_c = str(row[2].value).strip() if row[2].value else ""
                val_f = row[5].value
                if val_c and len(val_c) == 9:
                    if isinstance(val_f, datetime.time):
                        # Excel lưu thời lượng mà thực chất là mm:ss dưới dạng HH:MM
                        # (ví dụ 01:08 = 1 phút 8 giây → datetime.time(1, 8, 0))
                        # Nên: phút thực = .hour, giây thực = .minute
                        total_secs = val_f.hour * 60 + val_f.minute
                        mins = total_secs // 60
                        secs = total_secs % 60
                        id_to_time[val_c] = f"00:{mins:02d}:{secs:02d}"
                    elif isinstance(val_f, str):
                        tl_str = val_f.replace(" AM", "").replace(" PM", "").strip()
                        parts = tl_str.split(":")
                        if len(parts) >= 2:
                            try:
                                # Trường hợp string cũng có thể là "H:MM" = phút:giây
                                total_secs = int(parts[0]) * 60 + int(parts[1])
                                id_to_time[val_c] = f"00:{(total_secs // 60):02d}:{(total_secs % 60):02d}"
                            except:
                                pass

            # 2. Xử lý LIST bằng thuật toán nội bộ thay vì AI
            self.log("Bóc tách danh sách tin chính từ file LIST (sử dụng thuật toán nội bộ thay vì AI để tránh lỗi 504 Timeout)...")
            danh_sach_tin = []
            ngay_phat = ""
            
            for row_idx, row in enumerate(ws_list.iter_rows(values_only=False), start=1):
                val_a = str(row[0].value).strip() if row[0].value else ""
                val_c = str(row[2].value).strip() if row[2].value else ""
                val_d = str(row[3].value).strip() if row[3].value else ""
                
                if row_idx <= 5 and not ngay_phat:
                    for cell in row:
                        c_val = str(cell.value) if cell.value else ""
                        match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', c_val)
                        if match:
                            d, m, y = match.groups()
                            ngay_phat = f"{y}{m.zfill(2)}{d.zfill(2)}"
                            break
                            
                if (val_a.startswith("24H-") or val_a.startswith("GAT24H-")) and val_d == "ONLINE" and len(val_c) == 9 and val_c.isdigit():
                    danh_sach_tin.append({
                        "id": val_c,
                        "ten_file": val_a,
                        "thoi_luong": id_to_time.get(val_c, "00:00:00")
                    })
                    
            if not ngay_phat:
                match_fn = re.search(r'\d{8}', list_files[0])
                if match_fn:
                    ngay_phat = match_fn.group(0)
                else:
                    ngay_phat = datetime.datetime.now().strftime("%Y%m%d")

            self.log(f"Đã tìm thấy {len(danh_sach_tin)} bản tin chính. Ngày phát sóng: {ngay_phat}")

            # 3. Parse NHUNG NGUOI THUC HIEN.rtf
            self.progress_var.set(20)
            ekip_file = os.path.join(input_d, "NHUNG NGUOI THUC HIEN.rtf")
            crew_data = {}
            if os.path.exists(ekip_file):
                self.log("Bóc tách thông tin ê-kíp sản xuất từ NHUNG NGUOI THUC HIEN.rtf...")
                with open(ekip_file, 'rb') as f:
                    rtf_raw = f.read().decode('utf-8', errors='replace')
                ekip_text = rtf_to_text(rtf_raw)
                
                model_crew = genai.GenerativeModel(model_name)
                prompt_crew = f"""
Trích xuất thông tin người đảm nhận các chức danh từ văn bản ê-kíp chương trình. Nếu chức danh không có tên người, trả về chuỗi rỗng "". Tên người VIẾT HOA.
Văn bản:
{ekip_text}
"""
                for attempt in range(3):
                    try:
                        res_crew = model_crew.generate_content(
                            prompt_crew,
                            generation_config=genai.GenerationConfig(
                                response_mime_type="application/json",
                                response_schema=CrewList
                            )
                        )
                        crew_data = json.loads(res_crew.text)
                        break
                    except Exception as e:
                        if attempt == 2:
                            self.log(f"Lỗi khi bóc tách ê-kíp sau 3 lần thử: {e}")
                            crew_data = {k: "" for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat", "trang_diem"]}
                        else:
                            self.log(f"Lỗi API (Ê-kíp), đang thử lại... (lần {attempt+1})")
                            time.sleep(2)
            else:
                self.log("Không tìm thấy file NHUNG NGUOI THUC HIEN.rtf, bỏ qua ê-kíp.")
                crew_data = {k: "" for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat", "trang_diem"]}

            # Fallback: Bổ sung từ các file RTF tiền tố nếu crew_data còn thiếu
            PREFIX_MAP = {
                "BGĐ ": "chiu_trach_nhiem",
                "BT ":  "bien_tap",
                "BD ":  "bien_dich",
                "MC ":  "hien_dan",
                "ĐD ":  "dao_dien",
                "KT ":  "ky_thuat",
            }
            # Kiểm tra xem có trường nào bị rỗng không
            missing_keys = [k for k in ["chiu_trach_nhiem", "bien_tap", "bien_dich", "hien_dan", "dao_dien", "ky_thuat"] if not crew_data.get(k, "").strip()]
            if missing_keys:
                self.log("Một số chức danh còn thiếu tên — đang tìm file RTF tiền tố bổ sung...")
                # Ánh xạ key -> tiền tố cần tìm
                KEY_TO_PREFIX = {v: k for k, v in PREFIX_MAP.items()}
                for fname in os.listdir(input_d):
                    if not fname.endswith(".rtf"):
                        continue
                    for prefix, field_key in PREFIX_MAP.items():
                        if fname.startswith(prefix) and field_key in missing_keys:
                            # Lấy phần sau tiền tố, bỏ đuôi .rtf
                            name_part = fname[len(prefix):][:-4].strip()
                            # Luôn nối các tên bằng " - " (thay mọi dấu phẩy)
                            name_part = re.sub(r"\s*,\s*", " - ", name_part)
                            crew_data[field_key] = name_part.upper()
                            missing_keys.remove(field_key)
                            self.log(f"  Bổ sung từ '{fname}': {field_key} = {crew_data[field_key]}")
                            break  # Không cần quét thêm prefix cho file này

            # 4. Parse từng file RTF tin chính
            self.log("Bắt đầu bóc tách nội dung chi tiết từng file kịch bản (RTF)...")
            chi_tiet_tin = [] # Chứa dict: id, a245, a500, a520_list
            model_rtf = genai.GenerativeModel(model_name)

            total_tin = len(danh_sach_tin)
            for idx, tin in enumerate(danh_sach_tin):
                id_tin = tin["id"]
                ten_tin = tin["ten_file"]
                
                # Tìm file rtf tương ứng (xử lý ký tự / bị thay đổi trong tên file)
                safe_name = ten_tin.replace("/", "")
                rtf_path = None
                for fname in os.listdir(input_d):
                    if fname.endswith(".rtf") and safe_name.lower() in fname.replace("/", "").lower():
                        rtf_path = os.path.join(input_d, fname)
                        break
                
                if not rtf_path:
                    # Fallback tìm kiếm gần đúng bằng prefix
                    for fname in os.listdir(input_d):
                        if fname.endswith(".rtf") and fname.startswith(ten_tin[:15]):
                            rtf_path = os.path.join(input_d, fname)
                            break

                if not rtf_path:
                    self.log(f"CẢNH BÁO: Không tìm thấy file RTF cho tin '{ten_tin}' (ID: {id_tin})")
                    chi_tiet_tin.append({"id": id_tin, "tieu_de": ten_tin.upper(), "nguoi_bien_dich": "", "noi_dung": []})
                    continue

                self.log(f"Đang bóc tách: {os.path.basename(rtf_path)}...")
                with open(rtf_path, 'rb') as f:
                    rtf_raw = f.read().decode('utf-8', errors='replace')
                news_text = rtf_to_text(rtf_raw)

                # Cắt bỏ nội dung từ dòng có 3 ký tự '=' liên tục trở lên đến cuối
                filtered_lines = []
                for line in news_text.split('\n'):
                    if re.search(r'={3,}', line):
                        break
                    filtered_lines.append(line)
                news_text = '\n'.join(filtered_lines)

                prompt_news = f"""
Trích xuất thông tin từ kịch bản bản tin sau.
1. tieu_de: Tiêu đề bản tin, thường được viết HOA toàn bộ (ví dụ: THIỆT HẠI DO ĐỘNG ĐẤT Ở PHILIPPINES TIẾP TỤC TĂNG). Nó thường nằm ở dòng thứ 1, 3, hoặc 5 của văn bản. Bỏ qua dòng chữ 'GẠT TG24H' và các tiêu đề tiếng Anh. Bắt buộc phải trích xuất được tiêu đề.
2. nguoi_bien_dich: Tên người biên dịch bản tin (tên người Việt), thường nằm trước tiêu đề chính. Nếu không có hoặc không rõ thì để chuỗi rỗng "", TUYỆT ĐỐI KHÔNG tự bịa ra tên.
3. noi_dung: Danh sách các đoạn văn bản cấu thành nội dung tin. Bao gồm các dòng phụ đề viết HOA và các đoạn lời đọc. Bỏ qua các dòng mã hình/video, bỏ qua tên tiếng Anh, bỏ qua ngày tháng. Giữ nguyên format viết hoa của phụ đề. Mỗi đoạn/câu là một phần tử trong mảng.

Văn bản:
{news_text}
"""
                news_parsed = {}
                for attempt in range(3):
                    try:
                        res_news = model_rtf.generate_content(
                            prompt_news,
                            generation_config=genai.GenerationConfig(
                                response_mime_type="application/json",
                                response_schema=RtfParseResult
                            )
                        )
                        news_parsed = json.loads(res_news.text)
                        # Kiểm tra noi_dung có thực sự có nội dung không
                        if news_parsed.get("noi_dung") and len(news_parsed["noi_dung"]) > 0:
                            break  # Thành công thật sự
                        else:
                            self.log(f"  ⚠ API trả về noi_dung rỗng cho {os.path.basename(rtf_path)} (lần {attempt+1}/3), thử lại...")
                            if attempt < 2:
                                time.sleep(2)
                    except Exception as ex:
                        if attempt == 2:
                            self.log(f"Lỗi khi gọi API bóc tách {os.path.basename(rtf_path)} sau 3 lần thử: {ex}")
                        else:
                            self.log(f"Lỗi API ({os.path.basename(rtf_path)}), đang thử lại... (lần {attempt+1})")
                            time.sleep(2)
                
                tieu_de_ai = news_parsed.get("tieu_de", "").strip()
                noi_dung_parsed = news_parsed.get("noi_dung", [])

                if not tieu_de_ai:
                    # Fallback: tìm dòng viết hoa đầu tiên dài hơn 10 ký tự, không chứa GẠT TG24H
                    for line in news_text.split('\n'):
                        line = line.strip()
                        if line.isupper() and len(line) > 10 and "GẠT TG24H" not in line and "HEADLINES" not in line:
                            tieu_de_ai = line
                            break

                # Fallback nội dung: nếu sau 3 lần AI vẫn trả noi_dung rỗng,
                # tự trích xuất toàn bộ văn bản phía dưới tiêu đề
                if not noi_dung_parsed:
                    self.log(f"  ⚠ CẢNH BÁO: AI không trả về nội dung cho '{os.path.basename(rtf_path)}' sau 3 lần thử.")
                    self.log(f"  → Chuyển sang trích xuất nội dung bằng thuật toán nội bộ (không dùng AI)...")
                    all_lines = [l.strip() for l in news_text.split('\n')]
                    # Tìm vị trí tiêu đề trong văn bản
                    title_idx = -1
                    if tieu_de_ai:
                        for li, line in enumerate(all_lines):
                            if tieu_de_ai in line:
                                title_idx = li
                                break
                    if title_idx == -1:
                        # Nếu không tìm được tiêu đề, tìm dòng viết hoa đầu tiên dài > 10
                        for li, line in enumerate(all_lines):
                            if line.isupper() and len(line) > 10 and "GẠT TG24H" not in line and "HEADLINES" not in line:
                                title_idx = li
                                break
                    # Lấy tất cả đoạn phía dưới tiêu đề
                    if title_idx >= 0:
                        content_lines = all_lines[title_idx + 1:]
                    else:
                        content_lines = all_lines  # Không tìm được tiêu đề → lấy toàn bộ
                    # Lọc bỏ dòng trống và dòng quá ngắn (<=2 ký tự)
                    noi_dung_parsed = [l for l in content_lines if len(l) > 2]
                    self.log(f"  ✓ Đã trích xuất được {len(noi_dung_parsed)} đoạn nội dung bằng thuật toán nội bộ.")

                chi_tiet_tin.append({
                    "id": id_tin,
                    "tieu_de": tieu_de_ai,
                    "nguoi_bien_dich": news_parsed.get("nguoi_bien_dich", ""),
                    "noi_dung": noi_dung_parsed
                })

                self.progress_var.set(20 + (idx + 1) / total_tin * 60) # Cập nhật progress 20 -> 80%

            # 5. Sinh Output 1: Import_SoLuoc
            self.log("Đang tạo file Output 1: Import_SoLuoc...")
            wb1 = openpyxl.Workbook()
            ws1 = wb1.active
            ws1.title = "Sheet1"
            headers1 = ["STT", "$a090", "$a245", "$n245", "$p245", "$b245", "$a246", "$a260", "$b260", "$c260", "$a300", "$c300", "$a306", "$a490", "$a500", "$t773", "$r773", "$r773", "$a911"]
            ws1.append(headers1)
            
            nam = ngay_phat[:4] if len(ngay_phat) >= 4 else str(datetime.datetime.now().year)
            thang = ngay_phat[4:6] if len(ngay_phat) >= 6 else str(datetime.datetime.now().month).zfill(2)
            ngay = ngay_phat[6:8] if len(ngay_phat) >= 8 else str(datetime.datetime.now().day).zfill(2)

            for idx, tin in enumerate(danh_sach_tin):
                # Match title
                tieu_de = ""
                for ct in chi_tiet_tin:
                    if ct["id"] == tin["id"]:
                        tieu_de = ct["tieu_de"]
                        break
                
                row_data = [
                    f"{(idx+1):02d}", # STT
                    tin["id"],        # $a090
                    tieu_de,          # $a245
                    "", "",           # $n245, $p245
                    f"Tin thế giới - bản tin 24g ngày {ngay}/{thang}/{nam}", # $b245
                    "",               # $a246
                    "Tp.HCM",         # $a260
                    "Trung tâm tin tức HTV", # $b260
                    nam,              # $c260
                    "File MXF",       # $a300
                    "",               # $c300
                    tin["thoi_luong"],# $a306
                    "",               # $a490
                    f"Tên file: {tin['id']}  {tin['ten_file']}", # $a500
                    "", "", "",       # $t773, $r773 x2
                    "Trung tâm Phát hình - Tư liệu HTV" # $a911
                ]
                ws1.append(row_data)
            
            apply_tnr_font(ws1)
            fn1 = os.path.join(output_d, f"Import_SoLuoc_TG24H_{ngay_phat}.xlsx")
            wb1.save(fn1)

            # 6. Sinh Output 2: Map_BanTinTG
            self.progress_var.set(85)
            self.log("Đang tạo file Output 2: Map_BanTinTG...")
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = "Sheet1"
            headers2 = ["$a090", "$a500", "$a505", "$a911"]
            ws2.append(headers2)

            # Build a500 column array
            a500_crew = [
                f"CHỊU TRÁCH NHIỆM NỘI DUNG: {crew_data.get('chiu_trach_nhiem', '')}",
                f"BIÊN TẬP: {crew_data.get('bien_tap', '')}",
                f"BIÊN DỊCH: {crew_data.get('bien_dich', '')}",
                f"HIỆN DẪN: {crew_data.get('hien_dan', '')}",
                f"ĐẠO DIỄN: {crew_data.get('dao_dien', '')}",
                f"KỸ THUẬT: {crew_data.get('ky_thuat', '')}",
                f"TRANG ĐIỂM: {crew_data.get('trang_diem', '')}",
                "Website: www.htv.com.vn/tin-tuc",
                "Fanpage: www.fb.com/htvtintuc",
                "Kênh Youtube: www.youtube.com/c/htvtintuc"
            ]

            # Lặp theo danh sách tin để đủ số dòng
            max_rows = max(len(a500_crew), len(danh_sach_tin))
            
            for i in range(max_rows):
                val_a500 = a500_crew[i] if i < len(a500_crew) else ""
                
                val_a505 = ""
                if i < len(danh_sach_tin):
                    tin = danh_sach_tin[i]
                    tieu_de = ""
                    for ct in chi_tiet_tin:
                        if ct["id"] == tin["id"]:
                            tieu_de = ct["tieu_de"]
                            break
                    val_a505 = f"{(i+1):02d} - {tieu_de}. Thời lượng: {tin['thoi_luong']}. ID: {tin['id']}"
                
                val_a911 = "Phạm Thị Đông" if i == 0 else ""
                
                ws2.append(["K303419", val_a500, val_a505, val_a911])
            
            apply_tnr_font(ws2)
            fn2 = os.path.join(output_d, f"Map_BanTinTG_24G_{ngay_phat}.xlsx")
            wb2.save(fn2)

            # 7. Sinh Output 3: Map_ChiTiet
            self.progress_var.set(95)
            self.log("Đang tạo file Output 3: Map_ChiTiet...")
            wb3 = openpyxl.Workbook()
            ws3 = wb3.active
            ws3.title = "24G"
            ws3.append(["$a090", "$a500", "$a520"])

            for idx, ct in enumerate(chi_tiet_tin):
                id_tin = ct["id"]
                nguoi_bd = ct["nguoi_bien_dich"]
                nd_list = ct["noi_dung"]
                
                val_a500_first = f"Biên dịch: {nguoi_bd}" if nguoi_bd else ""
                
                if not nd_list:
                    # Nếu lỗi không bóc được nội dung, để trống 1 dòng
                    ws3.append([id_tin, val_a500_first, ""])
                else:
                    for j, line in enumerate(nd_list):
                        a500 = val_a500_first if j == 0 else ""
                        ws3.append([id_tin, a500, line])

            apply_tnr_font(ws3)
            fn3 = os.path.join(output_d, f"Map_ChiTiet_24G_{ngay_phat}.xlsx")
            wb3.save(fn3)

            self.progress_var.set(100)
            self.log("=== HOÀN TẤT BIÊN MỤC ===")
            messagebox.showinfo("Thành công", f"Đã sinh 3 file output thành công tại:\n{output_d}")

        except Exception as e:
            err_msg = traceback.format_exc()
            self.log(f"LỖI NGHIÊM TRỌNG: {e}\n{err_msg}")
            messagebox.showerror("Lỗi", f"Đã xảy ra lỗi, vui lòng xem log tiến trình trên màn hình!\nChi tiết: {e}")
        finally:
            self.btn_start.config(state="normal")

if __name__ == "__main__":
    root = tk.Tk()
    app = BienMucApp(root)
    root.mainloop()
